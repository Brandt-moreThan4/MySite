"""I want to make an easy interface for importing excel files into my models. 
Mainly just the knowledge repo for now."""

import openpyxl
from ..models import Knowledge, Book, BlogPost


IMPORT_PATH = r'C:\Users\15314\OneDrive\Desktop\MySite\MySite\blog\data\data_dump\websiteDB.xlsx'


def import_knowledge():
    """Take knowledge rec's from excel spreadsheet and upload the to sql database"""
    
    #add ability for it to only update records that are not already in there

    wb = openpyxl.load_workbook(IMPORT_PATH)
    sht = wb['Knowledge']
    
    for i in range(2, get_last_row(sht, 'B') + 1):
        new_knowledge = Knowledge()
        new_knowledge.author = sht["B" + str(i)].value
        new_knowledge.description = sht["C" + str(i)].value
        new_knowledge.source = sht["D" + str(i)].value
        new_knowledge.tags = sht["E" + str(i)].value
        new_knowledge.save()



def update_knowledge():
    """This updates the knowledge table keywords based on what's in the spreadsheet. Uses primary key to identify which record to update."""
    wb = openpyxl.load_workbook(IMPORT_PATH)
    sht = wb['Knowledge']

    for i in range(2, get_last_row(sht, 'B') + 1):
        print(f'I am on excel row: {i} about to alter: {sht["C" + str(i)].value}')
        id = sht["A" + str(i)].value
        current_knowledge = Knowledge.objects.get(pk=sht["A" + str(i)].value)
        taggies = sht["E" + str(i)].value
        current_knowledge.tags = taggies
        current_knowledge.save()

def import_books():
    """Take books from excel spreadsheet and upload the to sql database"""
    
    #add ability for it to only update records that are not already in there

    wb = openpyxl.load_workbook(IMPORT_PATH)
    sht = wb['Books']

    for i in range(2, get_last_row(sht, 'B') + 1):
        new_book = Book()
        new_book.book_title = sht["B" + str(i)].value
        new_book.slug = sht["C" + str(i)].value
        new_book.author = sht["D" + str(i)].value
        new_book.cover_description = sht["E" + str(i)].value
        new_book.post_body = sht["F" + str(i)].value
        new_book.image_name = sht["G" + str(i)].value       
        #new_book.created = sht["H" + str(i)].value 
        new_book.save()


def update_books():
    """This updates the book recs based on primary key"""
    wb = openpyxl.load_workbook(IMPORT_PATH)
    sht = wb['Books']

    for i in range(2, get_last_row(sht, 'A') + 1):
        print(f'I am on excel row: {i} about to alter: {sht["C" + str(i)].value}')
        id = sht["A" + str(i)].value
        current_book = Book.objects.get(pk=id)      
        current_book.author = sht["G" + str(i)].value
        current_book.save()



def import_blog():
    """Take books from excel spreadsheet and upload the to sql database"""
    
    #add ability for it to only update records that are not already in there

    wb = openpyxl.load_workbook(IMPORT_PATH)
    sht = wb['Blog']

    for i in range(2, get_last_row(sht, 'B') + 1):
        new_blog = BlogPost()
        new_blog.post_title = sht["B" + str(i)].value
        new_blog.slug = sht["C" + str(i)].value
        new_blog.post_body = sht["D" + str(i)].value
        new_blog.save()



def get_last_row(sht, column='A'):
    """Return the last row with data in a worksheet for that column
    This should actually start from the bottom and come up but I am not
    sure how to work backwars other than using some arbitrary big number"""
    
    row = 1
    while sht[f'{column}{row}'].value is not None:
        row += 1
    return row - 1


